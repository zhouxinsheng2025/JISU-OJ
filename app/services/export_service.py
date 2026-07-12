"""Excel 成绩导出服务"""
import io
from datetime import datetime
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from app.models import Contest, ContestProblem, Problem, User, ScoreboardCache, ScoreMode


async def export_contest_excel(db: AsyncSession, contest_id: int) -> bytes:
    """导出比赛成绩为 Excel (.xlsx)"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "比赛成绩"

    # 获取比赛信息
    contest = (await db.execute(select(Contest).where(Contest.id == contest_id))).scalar_one_or_none()
    if not contest:
        return b""

    # 获取题目列表
    cp_result = await db.execute(
        select(Problem)
        .join(ContestProblem, ContestProblem.problem_id == Problem.id)
        .where(ContestProblem.contest_id == contest_id)
        .order_by(ContestProblem.order)
    )
    problems = list(cp_result.scalars().all())

    # 获取成绩数据
    board_result = await db.execute(
        select(ScoreboardCache).where(ScoreboardCache.contest_id == contest_id)
    )
    entries = list(board_result.scalars().all())

    # 按队伍分组统计
    team_stats = {}
    for e in entries:
        if e.team_id not in team_stats:
            team_stats[e.team_id] = {"solved": 0, "total_time": 0, "total_score": 0.0, "problems": {}}
        team_stats[e.team_id]["problems"][e.problem_id] = e
        if e.is_correct:
            team_stats[e.team_id]["solved"] += 1
            team_stats[e.team_id]["total_time"] += e.total_time
        team_stats[e.team_id]["total_score"] += e.score

    # 获取队伍名
    team_ids = list(team_stats.keys())
    users = {}
    if team_ids:
        user_result = await db.execute(select(User).where(User.id.in_(team_ids)))
        users = {u.id: u for u in user_result.scalars().all()}

    # 排序
    if contest.score_mode == ScoreMode.IOI:
        sorted_teams = sorted(team_stats.items(), key=lambda x: (-x[1]["total_score"], x[1]["total_time"]))
    else:
        sorted_teams = sorted(team_stats.items(), key=lambda x: (-x[1]["solved"], x[1]["total_time"]))

    # 样式
    header_font = Font(name='Microsoft YaHei', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    ac_fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
    wa_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    gold_fill = PatternFill(start_color='FEF9C3', end_color='FEF9C3', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )
    center = Alignment(horizontal='center', vertical='center')

    # 标题行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4 + len(problems))
    title_cell = ws.cell(row=1, column=1, value=f"{contest.title} — 比赛成绩")
    title_cell.font = Font(name='Microsoft YaHei', bold=True, size=14, color='0F2440')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # 信息行
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4 + len(problems))
    info_cell = ws.cell(row=2, column=1,
                        value=f"计分模式: {contest.score_mode.value.upper()}  |  "
                              f"时间: {contest.start_time.strftime('%Y-%m-%d %H:%M')} ~ {contest.end_time.strftime('%Y-%m-%d %H:%M')}  |  "
                              f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    info_cell.font = Font(name='Microsoft YaHei', size=9, color='6B7280')
    info_cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 22

    # 表头
    headers = ['排名', '队伍'] + [p.title for p in problems] + ['解题数', '罚时/总分']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
    ws.row_dimensions[4].height = 24

    # 数据行
    for rank, (team_id, stats) in enumerate(sorted_teams, 1):
        row = rank + 4
        user = users.get(team_id)
        teamname = user.teamname if user else f"Team {team_id}"

        ws.cell(row=row, column=1, value=rank).alignment = center
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=2, value=teamname).border = thin_border
        ws.cell(row=row, column=2).font = Font(name='Microsoft YaHei', bold=(rank <= 3))

        for pi, p in enumerate(problems):
            entry = stats["problems"].get(p.id)
            cell = ws.cell(row=row, column=3 + pi)
            cell.alignment = center
            cell.border = thin_border
            if entry:
                if entry.is_correct:
                    cell.value = f"AC ({entry.submissions})"
                    cell.fill = ac_fill
                    cell.font = Font(color='16A34A', bold=True)
                else:
                    cell.value = f"WA ({entry.submissions})"
                    cell.fill = wa_fill
                    cell.font = Font(color='DC2626')
            else:
                cell.value = "-"
                cell.font = Font(color='9CA3AF')

        solved_cell = ws.cell(row=row, column=3 + len(problems), value=stats["solved"])
        solved_cell.alignment = center
        solved_cell.border = thin_border
        solved_cell.font = Font(bold=True, color='D4A843')

        total_cell = ws.cell(row=row, column=4 + len(problems),
                             value=round(stats["total_score"], 2) if contest.score_mode == ScoreMode.IOI else stats["total_time"])
        total_cell.alignment = center
        total_cell.border = thin_border

        # 前三名高亮
        if rank <= 3:
            for c in range(1, 4 + len(problems)):
                if ws.cell(row=row, column=c).fill == PatternFill():
                    ws.cell(row=row, column=c).fill = gold_fill

    # 列宽
    from openpyxl.utils import get_column_letter
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 18
    for pi in range(len(problems)):
        col_letter = get_column_letter(3 + pi)  # C=3, D=4, ...
        ws.column_dimensions[col_letter].width = 14
    ws.column_dimensions[get_column_letter(3 + len(problems))].width = 10
    ws.column_dimensions[get_column_letter(4 + len(problems))].width = 12

    # 冻结表头
    ws.freeze_panes = 'A5'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
